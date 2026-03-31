"""Seed the BankOfferAI database from the Excel hackathon dataset."""

import json
import logging
import os
import sys

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://postgres:postgres@localhost:5432/bankofferingai",
)

# Resolve the Excel file relative to this script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "..", "AI_Hackathon_Product_Offering_Engine_Dataset_v1.xlsx")


def read_sheet(wb, sheet_name, header_row=0):
    """Read a sheet into a list of dicts using the given header row."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
               for i, h in enumerate(rows[header_row])]
    data = []
    for row in rows[header_row + 1:]:
        if all(v is None for v in row[:len(headers)]):
            continue
        record = {}
        for i, h in enumerate(headers):
            record[h] = row[i] if i < len(row) else None
        data.append(record)
    return data


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def run_schema(conn):
    """Execute schema.sql to create tables."""
    schema_path = os.path.join(SCRIPT_DIR, "schema.sql")
    with open(schema_path) as f:
        sql = f.read()
    with conn.cursor() as cur:
        cur.execute(sql)
    conn.commit()
    logger.info("Schema applied")


def seed_customers(conn, customers_data):
    with conn.cursor() as cur:
        for c in customers_data:
            cid = str(c.get("customer_id", ""))
            if not cid:
                continue
            # Header has "dependents_count_(kids)"
            deps_key = next((k for k in c if "dependents" in k), "dependents_count")
            cur.execute("""
                INSERT INTO customers (customer_id, age, city, income, savings, debt,
                    has_debt, risk_profile, marital_status, dependents_count,
                    homeowner_status, existing_products)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (customer_id) DO NOTHING
            """, (
                cid,
                safe_int(c.get("age")),
                c.get("city"),
                safe_float(c.get("income")),
                safe_float(c.get("savings")),
                safe_float(c.get("debt")),
                bool(safe_int(c.get("has_debt"))),
                c.get("risk_profile"),
                c.get("marital_status"),
                safe_int(c.get(deps_key)),
                c.get("homeowner_status"),
                c.get("existing_products"),
            ))
    conn.commit()
    logger.info("Seeded %d customers", len(customers_data))


def seed_customer_features(conn, customers_data, features_data):
    """Merge customers_enhanced + features_enhanced into customer_features."""
    features_by_id = {}
    for f in features_data:
        cid = str(f.get("customer_id", ""))
        if cid:
            features_by_id[cid] = f

    with conn.cursor() as cur:
        for c in customers_data:
            cid = str(c.get("customer_id", ""))
            if not cid:
                continue
            feat = features_by_id.get(cid, {})
            deps_key = next((k for k in c if "dependents" in k), "dependents_count")
            monthly_income = safe_float(c.get("income"))
            annual_income = monthly_income * 12

            # Clamp savings_rate and debt_to_income to reasonable ranges
            raw_savings_rate = safe_float(feat.get("savings_rate", 0))
            savings_rate = max(0.0, min(1.0, raw_savings_rate))

            raw_dti = safe_float(feat.get("debt_to_income", 0))
            debt_to_income = max(0.0, min(10.0, raw_dti))

            cur.execute("""
                INSERT INTO customer_features (
                    customer_id, age, annual_income, dependents, account_tenure_years,
                    investment_balance, savings_rate, debt_to_income,
                    monthly_savings, avg_expenses, idle_cash, balance_trend,
                    dominant_spend_category, investment_gap_flag,
                    risk_profile, city, marital_status, homeowner_status, existing_products
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (customer_id) DO NOTHING
            """, (
                cid,
                safe_int(c.get("age")),
                annual_income,
                safe_int(c.get(deps_key)),
                0,  # account_tenure_years not in dataset
                safe_float(c.get("savings")),  # approximate investment_balance from savings
                savings_rate,
                debt_to_income,
                safe_float(feat.get("monthly_savings")),
                safe_float(feat.get("avg_expenses")),
                safe_float(feat.get("idle_cash")),
                feat.get("balance_trend"),
                feat.get("dominant_spend_category"),
                safe_int(feat.get("investment_gap_flag")),
                c.get("risk_profile"),
                c.get("city"),
                c.get("marital_status"),
                c.get("homeowner_status"),
                c.get("existing_products"),
            ))
    conn.commit()
    logger.info("Seeded %d customer_features rows", len(customers_data))


def seed_transactions(conn, txn_data):
    with conn.cursor() as cur:
        for t in txn_data:
            cid = str(t.get("customer_id", ""))
            if not cid:
                continue
            cur.execute("""
                INSERT INTO transactions (customer_id, transaction_date, amount, category, channel, recurring_flag)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (
                cid,
                t.get("date"),
                safe_float(t.get("amount")),
                t.get("category"),
                t.get("channel"),
                bool(safe_int(t.get("recurring_flag"))),
            ))
    conn.commit()
    logger.info("Seeded %d transactions", len(txn_data))


def seed_events(conn, events_data):
    with conn.cursor() as cur:
        for e in events_data:
            cid = str(e.get("customer_id", ""))
            if not cid:
                continue
            cur.execute("""
                INSERT INTO events (customer_id, event_type, event_date)
                VALUES (%s,%s,%s)
            """, (
                cid,
                e.get("event_type"),
                e.get("date"),
            ))
    conn.commit()
    logger.info("Seeded %d events", len(events_data))


def seed_products(conn, product_data):
    with conn.cursor() as cur:
        for p in product_data:
            name = p.get("product_name")
            if not name:
                continue
            cur.execute("""
                INSERT INTO products (product_name, category, short_description, channel, priority, lifecycle_stage, when_to_recommend)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (product_name) DO NOTHING
            """, (
                name,
                p.get("category"),
                p.get("short_description"),
                p.get("channel"),
                p.get("priority"),
                p.get("lifecycle_stage"),
                p.get("when_to_recommend"),
            ))
    conn.commit()
    logger.info("Seeded %d products", len(product_data))


def seed_profiles(conn, customers_data, features_data):
    """Pre-build customer profiles and store as JSON for the profiles endpoint."""
    # Add worker modules to path
    sys.path.insert(0, os.path.join(SCRIPT_DIR, ".."))

    from services.worker.profiler import build_profile

    features_by_id = {str(f.get("customer_id", "")): f for f in features_data}

    with conn.cursor() as cur:
        for c in customers_data:
            cid = str(c.get("customer_id", ""))
            if not cid:
                continue
            feat = features_by_id.get(cid, {})
            deps_key = next((k for k in c if "dependents" in k), "dependents_count")
            monthly_income = safe_float(c.get("income"))

            raw_sr = safe_float(feat.get("savings_rate", 0))
            raw_dti = safe_float(feat.get("debt_to_income", 0))

            profiler_input = {
                "age": safe_int(c.get("age"), 30),
                "annual_income": max(1.0, monthly_income * 12),
                "dependents": safe_int(c.get(deps_key)),
                "account_tenure_years": 0,
                "investment_balance": safe_float(c.get("savings")),
                "savings_ratio": max(0.0, min(1.0, raw_sr)),
                "loan_to_income": max(0.0, min(1.0, raw_dti)),
            }

            try:
                profile = build_profile(cid, profiler_input)
                profile_json = json.dumps({
                    "customer_id": cid,
                    "life_stage": profile.life_stage,
                    "risk_score": profile.risk_score,
                    "segments": profile.segments,
                    "income_bracket": (
                        "very_high" if monthly_income * 12 >= 150000 else
                        "high" if monthly_income * 12 >= 80000 else
                        "medium" if monthly_income * 12 >= 40000 else
                        "low"
                    ),
                    "spending_patterns": [],
                })
                cur.execute("""
                    INSERT INTO customer_profiles (customer_id, data)
                    VALUES (%s, %s)
                    ON CONFLICT (customer_id) DO UPDATE SET data = EXCLUDED.data
                """, (cid, profile_json))
            except Exception as e:
                logger.warning("Failed to build profile for %s: %s", cid, e)

    conn.commit()
    logger.info("Seeded customer profiles")


def main():
    logger.info("Loading Excel dataset from %s", XLSX_PATH)
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)

    customers_data = read_sheet(wb, "customers_enhanced")
    features_data = read_sheet(wb, "features_enhanced")
    txn_data = read_sheet(wb, "transactions_enhanced")
    events_data = read_sheet(wb, "events")
    product_data = read_sheet(wb, "Product Catalog")

    wb.close()
    logger.info("Loaded: %d customers, %d features, %d transactions, %d events, %d products",
                len(customers_data), len(features_data), len(txn_data),
                len(events_data), len(product_data))

    # Parse DATABASE_URL for psycopg2 (strip asyncpg scheme if present)
    db_url = DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://")
    logger.info("Connecting to database...")
    conn = psycopg2.connect(db_url)

    try:
        run_schema(conn)
        seed_customers(conn, customers_data)
        seed_customer_features(conn, customers_data, features_data)
        seed_transactions(conn, txn_data)
        seed_events(conn, events_data)
        seed_products(conn, product_data)
        seed_profiles(conn, customers_data, features_data)
        logger.info("Seeding complete!")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
